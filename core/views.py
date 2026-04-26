from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import IntegrityError
from django.db.models import Count, Q
from django.utils import timezone
from functools import wraps
import json
from datetime import date

from .models import CustomUser, Class, Section, Subject, Teacher, Student, Attendance, AttendanceRecord
from .forms import (
    LoginForm, ClassForm, SectionForm, SubjectForm, 
    TeacherCreationForm, TeacherAssignmentForm, StudentForm, AttendanceFilterForm, AttendanceExportForm,
    TeacherAllocationForm, ManagerCreationForm
)


# Custom decorators for role-based access control
def admin_required(view_func):
    """Decorator to ensure user is an admin"""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if request.user.role != 'ADMIN':
            messages.error(request, 'You do not have permission to access this page.')
            return redirect('teacher_dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


def teacher_required(view_func):
    """Decorator to ensure user is a teacher"""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if request.user.role != 'TEACHER':
            messages.error(request, 'You do not have permission to access this page.')
            return redirect('admin_dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


def manager_required(view_func):
    """Decorator to ensure user is a manager"""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if request.user.role != 'MANAGER':
            messages.error(request, 'You do not have permission to access this page.')
            # Redirect to admin if admin, teacher if teacher
            if request.user.role == 'ADMIN':
                return redirect('admin_dashboard')
            else:
                return redirect('teacher_dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


# Authentication Views
def login_view(request):
    """Handle user login"""
    if request.user.is_authenticated:
        if request.user.role == 'ADMIN':
            return redirect('admin_dashboard')
        elif request.user.role == 'MANAGER':
            return redirect('manager_dashboard')
        return redirect('teacher_dashboard')
    
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            
            if user is not None:
                login(request, user)
                messages.success(request, f'Welcome back, {user.get_full_name() or user.username}!')
                
                # Redirect based on role
                if user.role == 'ADMIN':
                    return redirect('admin_dashboard')
                elif user.role == 'MANAGER':
                    return redirect('manager_dashboard')
                else:
                    return redirect('teacher_dashboard')
            else:
                messages.error(request, 'Invalid username or password.')
    else:
        form = LoginForm()
    
    return render(request, 'login.html', {'form': form})


@login_required
def logout_view(request):
    """Handle user logout"""
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('login')


# Admin Views
@admin_required
def admin_dashboard(request):
    """Admin dashboard with statistics"""
    context = {
        'total_classes': Class.objects.count(),
        'total_subjects': Subject.objects.count(),
        'total_teachers': Teacher.objects.count(),
        'total_students': Student.objects.count(),
        'total_managers': CustomUser.objects.filter(role='MANAGER').count(),
    }
    return render(request, 'admin/dashboard.html', context)


@admin_required
def manage_managers(request):
    """Manage managers"""
    form = ManagerCreationForm()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_manager':
            form = ManagerCreationForm(request.POST)
            if form.is_valid():
                user = form.save(commit=False)
                user.role = 'MANAGER'
                user.save()
                messages.success(request, f'Manager {user.get_full_name()} created successfully.')
                return redirect('manage_managers')
            else:
                messages.error(request, 'Please correct the errors below.')
        
        elif action == 'edit_credentials':
            manager_id = request.POST.get('manager_id')
            new_username = request.POST.get('username')
            new_password = request.POST.get('password')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            
            try:
                user = CustomUser.objects.get(id=manager_id, role='MANAGER')
                
                if new_username and new_username != user.username:
                    if CustomUser.objects.filter(username=new_username).exists():
                        messages.error(request, 'Username already exists. Please choose another.')
                        return redirect('manage_managers')
                    user.username = new_username
                
                if first_name is not None:
                    user.first_name = first_name
                if last_name is not None:
                    user.last_name = last_name
                
                if new_password:
                    user.set_password(new_password)
                
                user.save()
                messages.success(request, f"Details updated for {user.get_full_name() or user.username}.")
            except CustomUser.DoesNotExist:
                messages.error(request, 'Manager not found.')
            return redirect('manage_managers')
            
        elif action == 'delete_manager':
            manager_id = request.POST.get('manager_id')
            try:
                user = CustomUser.objects.get(id=manager_id, role='MANAGER')
                user.delete()
                messages.success(request, 'Manager deleted successfully.')
            except CustomUser.DoesNotExist:
                messages.error(request, 'Manager not found.')
            return redirect('manage_managers')
            
    managers = CustomUser.objects.filter(role='MANAGER')
    
    context = {
        'form': form,
        'managers': managers,
    }
    return render(request, 'admin/manage_managers.html', context)


@admin_required
def manage_classes(request):
    """Manage classes and sections"""
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_class':
            form = ClassForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'Class created successfully.')
                return redirect('manage_classes')
        
        elif action == 'add_section':
            form = SectionForm(request.POST)
            if form.is_valid():
                try:
                    form.save()
                    messages.success(request, 'Section created successfully.')
                except IntegrityError:
                    messages.error(request, 'This section already exists for the selected class.')
                return redirect('manage_classes')
        
        elif action == 'delete_class':
            class_id = request.POST.get('class_id')
            try:
                class_obj = Class.objects.get(id=class_id)
                class_obj.delete()
                messages.success(request, 'Class deleted successfully.')
            except Class.DoesNotExist:
                messages.error(request, 'Class not found.')
            return redirect('manage_classes')
        
        elif action == 'delete_section':
            section_id = request.POST.get('section_id')
            try:
                section = Section.objects.get(id=section_id)
                section.delete()
                messages.success(request, 'Section deleted successfully.')
            except Section.DoesNotExist:
                messages.error(request, 'Section not found.')
            return redirect('manage_classes')
    
    class_form = ClassForm()
    section_form = SectionForm()
    classes = Class.objects.prefetch_related('sections').all()
    
    context = {
        'class_form': class_form,
        'section_form': section_form,
        'classes': classes,
    }
    return render(request, 'admin/manage_classes.html', context)


@admin_required
def manage_subjects(request):
    """Manage subjects"""
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_subject':
            form = SubjectForm(request.POST)
            if form.is_valid():
                try:
                    form.save()
                    messages.success(request, 'Subject created successfully.')
                except IntegrityError:
                    messages.error(request, 'This subject already exists for the selected class.')
                return redirect('manage_subjects')
        
        elif action == 'delete_subject':
            subject_id = request.POST.get('subject_id')
            try:
                subject = Subject.objects.get(id=subject_id)
                subject.delete()
                messages.success(request, 'Subject deleted successfully.')
            except Subject.DoesNotExist:
                messages.error(request, 'Subject not found.')
            return redirect('manage_subjects')
    
    subject_form = SubjectForm()
    subjects = Subject.objects.select_related('class_ref', 'section').all()
    
    context = {
        'subject_form': subject_form,
        'subjects': subjects,
    }
    return render(request, 'admin/manage_subjects.html', context)


@admin_required
def manage_teachers(request):
    """Manage teachers and their assignments"""
    teacher_form = TeacherCreationForm()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_teacher':
            teacher_form = TeacherCreationForm(request.POST)
            if teacher_form.is_valid():
                try:
                    # Create user with TEACHER role
                    user = teacher_form.save(commit=False)
                    user.role = 'TEACHER'
                    user.save()
                    
                    # Create teacher profile
                    Teacher.objects.create(
                        user=user,
                        employee_id=teacher_form.cleaned_data['employee_id'],
                        phone=teacher_form.cleaned_data.get('phone', ''),
                        password_text=teacher_form.cleaned_data['password1']
                    )
                    
                    messages.success(request, f'Teacher {user.get_full_name()} created successfully.')
                    return redirect('manage_teachers')
                except IntegrityError:
                    messages.error(request, 'Employee ID already exists.')
            else:
                # Form is invalid, show errors
                messages.error(request, 'Please correct the errors below.')
        
        elif action == 'assign':
            teacher_id = request.POST.get('teacher_id')
            teacher = get_object_or_404(Teacher, id=teacher_id)
            form = TeacherAssignmentForm(request.POST, instance=teacher)
            if form.is_valid():
                form.save()
                messages.success(request, 'Assignments updated successfully.')
                return redirect('manage_teachers')
        
        elif action == 'edit_credentials':
            teacher_id = request.POST.get('teacher_id')
            new_username = request.POST.get('username')
            new_password = request.POST.get('password')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            employee_id = request.POST.get('employee_id')
            phone = request.POST.get('phone')
            
            try:
                teacher = Teacher.objects.get(id=teacher_id)
                user = teacher.user
                
                # Update username if provided and changed
                if new_username and new_username != user.username:
                    # Check if new username already exists
                    if CustomUser.objects.filter(username=new_username).exists():
                        messages.error(request, 'Username already exists. Please choose another.')
                        return redirect('manage_teachers')
                    user.username = new_username
                
                if first_name is not None:
                    user.first_name = first_name
                if last_name is not None:
                    user.last_name = last_name
                if employee_id is not None:
                    teacher.employee_id = employee_id
                if phone is not None:
                    teacher.phone = phone
                
                # Update password if provided
                if new_password:
                    user.set_password(new_password)
                    teacher.password_text = new_password
                
                user.save()
                teacher.save()
                messages.success(request, f"Details updated for {user.get_full_name() or user.username}.")
            except Teacher.DoesNotExist:
                messages.error(request, 'Teacher not found.')
            return redirect('manage_teachers')
        
        elif action == 'delete_teacher':
            teacher_id = request.POST.get('teacher_id')
            try:
                teacher = Teacher.objects.get(id=teacher_id)
                user = teacher.user
                teacher.delete()
                user.delete()
                messages.success(request, 'Teacher deleted successfully.')
            except Teacher.DoesNotExist:
                messages.error(request, 'Teacher not found.')
            return redirect('manage_teachers')
            
        elif action == 'remove_subject':
            teacher_id = request.POST.get('teacher_id')
            subject_id = request.POST.get('subject_id')
            try:
                teacher = Teacher.objects.get(id=teacher_id)
                subject = Subject.objects.get(id=subject_id)
                teacher.assigned_subjects.remove(subject)
                messages.success(request, f'Removed {subject.name} from {teacher.user.get_full_name() or teacher.user.username}.')
            except (Teacher.DoesNotExist, Subject.DoesNotExist):
                messages.error(request, 'Teacher or Subject not found.')
            return redirect('manage_teachers')
    
    teachers = Teacher.objects.select_related('user').prefetch_related('assigned_subjects', 'assigned_classes').all()
    
    context = {
        'teacher_form': teacher_form,
        'teachers': teachers,
    }
    return render(request, 'admin/manage_teachers.html', context)


@admin_required
def manage_students(request):
    """Manage students"""
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_student':
            form = StudentForm(request.POST)
            if form.is_valid():
                try:
                    form.save()
                    messages.success(request, 'Student added successfully.')
                except IntegrityError:
                    messages.error(request, 'Roll number already exists.')
                return redirect('manage_students')
                
        elif action == 'edit_student':
            student_id = request.POST.get('student_id')
            try:
                student = Student.objects.get(id=student_id)
                student.roll_number = request.POST.get('roll_number')
                student.name = request.POST.get('name')
                
                section_id = request.POST.get('section')
                if section_id:
                    student.section = Section.objects.get(id=section_id)
                
                student.email = request.POST.get('email', '')
                student.phone = request.POST.get('phone', '')
                
                student.save()
                messages.success(request, 'Student updated successfully.')
            except Student.DoesNotExist:
                messages.error(request, 'Student not found.')
            except IntegrityError:
                messages.error(request, 'Error: Roll number already exists for another student.')
            except Section.DoesNotExist:
                messages.error(request, 'Selected section not found.')
            return redirect('manage_students')
        
        elif action == 'delete_student':
            student_id = request.POST.get('student_id')
            try:
                student = Student.objects.get(id=student_id)
                student.delete()
                messages.success(request, 'Student deleted successfully.')
            except Student.DoesNotExist:
                messages.error(request, 'Student not found.')
            return redirect('manage_students')
    
    # Get filter parameters from URL
    class_filter = request.GET.get('class_filter', '')
    section_filter = request.GET.get('section_filter', '')
    
    try:
        class_filter_id = int(class_filter)
    except ValueError:
        class_filter_id = None
        
    try:
        section_filter_id = int(section_filter)
    except ValueError:
        section_filter_id = None
    
    # Start with all students, ordered alphabetically by name
    students = Student.objects.select_related('section', 'section__class_ref').order_by('name')
    
    # Apply filters if provided
    if class_filter:
        students = students.filter(section__class_ref_id=class_filter)
    if section_filter:
        students = students.filter(section_id=section_filter)
    
    # Get all classes and sections for filter dropdowns
    classes = Class.objects.all().order_by('name')
    sections = Section.objects.select_related('class_ref').order_by('class_ref__name', 'name')
    
    student_form = StudentForm()
    
    context = {
        'student_form': student_form,
        'students': students,
        'classes': classes,
        'sections': sections,
        'class_filter': class_filter,
        'section_filter': section_filter,
        'class_filter_id': class_filter_id,
        'section_filter_id': section_filter_id,
    }
    return render(request, 'admin/manage_students.html', context)

@admin_required
def manage_allocations(request):
    """Manage teacher allocations (Classes and Subjects)"""
    if request.method == 'POST':
        form = TeacherAllocationForm(request.POST)
        if form.is_valid():
            teacher = form.cleaned_data['teacher']
            assigned_classes = form.cleaned_data['assigned_classes']
            assigned_sections = form.cleaned_data['assigned_sections']
            assigned_subjects = form.cleaned_data['assigned_subjects']
            
            # Make it ADDITIVE
            teacher.assigned_classes.add(*assigned_classes)
            teacher.assigned_sections.add(*assigned_sections)
            teacher.assigned_subjects.add(*assigned_subjects)
            
            messages.success(request, f'Allocations updated for {teacher.user.get_full_name()}.')
            return redirect('manage_allocations')
    else:
        form = TeacherAllocationForm()
    
    teachers = Teacher.objects.select_related('user').prefetch_related('assigned_subjects', 'assigned_classes').all()
    
    context = {
        'form': form,
        'teachers': teachers,
    }
    return render(request, 'admin/manage_allocations.html', context)


@admin_required
def export_students_excel(request):
    """Export filtered students to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    from datetime import datetime
    
    # Get filter parameters
    class_filter = request.GET.get('class_filter', '')
    section_filter = request.GET.get('section_filter', '')
    
    # Get filtered students
    students = Student.objects.select_related('section', 'section__class_ref').order_by('name')
    if class_filter:
        students = students.filter(section__class_ref_id=class_filter)
    if section_filter:
        students = students.filter(section_id=section_filter)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Add title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "Student List Report"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Add filter info
    filter_info = "Filters: "
    if class_filter:
        class_obj = Class.objects.get(id=class_filter)
        filter_info += f"Class: {class_obj.name} "
    if section_filter:
        section_obj = Section.objects.get(id=section_filter)
        filter_info += f"Section: {section_obj.name}"
    if not class_filter and not section_filter:
        filter_info += "All Students"
    
    ws.merge_cells('A2:F2')
    ws['A2'] = filter_info
    ws['A2'].font = Font(italic=True)
    
    # Add headers
    headers = ['Roll Number', 'Name', 'Class', 'Section', 'Email', 'Phone']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_num, student in enumerate(students, 5):
        ws.cell(row=row_num, column=1, value=student.roll_number)
        ws.cell(row=row_num, column=2, value=student.name)
        ws.cell(row=row_num, column=3, value=student.section.class_ref.name)
        ws.cell(row=row_num, column=4, value=student.section.name)
        ws.cell(row=row_num, column=5, value=student.email or '-')
        ws.cell(row=row_num, column=6, value=student.phone or '-')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"students_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@admin_required
def export_students_pdf(request):
    """Export filtered students to PDF"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from django.http import HttpResponse
    from datetime import datetime
    
    # Get filter parameters
    class_filter = request.GET.get('class_filter', '')
    section_filter = request.GET.get('section_filter', '')
    
    # Get filtered students
    students = Student.objects.select_related('section', 'section__class_ref').order_by('name')
    if class_filter:
        students = students.filter(section__class_ref_id=class_filter)
    if section_filter:
        students = students.filter(section_id=section_filter)
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    filename = f"students_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Create PDF
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Add title
    title = Paragraph("Student List Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Add filter info
    filter_info = "Filters: "
    if class_filter:
        class_obj = Class.objects.get(id=class_filter)
        filter_info += f"Class: {class_obj.name} "
    if section_filter:
        section_obj = Section.objects.get(id=section_filter)
        filter_info += f"Section: {section_obj.name}"
    if not class_filter and not section_filter:
        filter_info += "All Students"
    
    filter_para = Paragraph(filter_info, styles['Normal'])
    elements.append(filter_para)
    elements.append(Spacer(1, 0.3*inch))
    
    # Create table data
    data = [['Roll No.', 'Name', 'Class', 'Section', 'Email']]
    for student in students:
        data.append([
            student.roll_number,
            student.name,
            student.section.class_ref.name,
            student.section.name,
            student.email or '-'
        ])
    
    # Create table
    table = Table(data, colWidths=[1*inch, 2*inch, 1.5*inch, 1*inch, 2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    return response


# Manager Views
@manager_required
def manager_dashboard(request):
    """Manager dashboard showing absentee report for a selected date"""
    selected_date_str = request.GET.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        selected_date = date.fromisoformat(selected_date_str)
    except ValueError:
        selected_date = date.today()
        selected_date_str = selected_date.strftime('%Y-%m-%d')
    
    # Get all classes, their sections, and their subjects (with teachers)
    classes = Class.objects.prefetch_related(
        'sections', 
        'subjects__teachers__user'
    ).all().order_by('name')
    
    # Get all attendance sessions for the selected date
    attendances_today = Attendance.objects.filter(date=selected_date)
    
    # Create a set of (section_id, subject_id) that have attendance marked today
    marked_sessions = set(attendances_today.values_list('section_id', 'subject_id'))
    
    # Group results by class, section
    grouped_data = {}
    total_missing_attendance = 0
    
    for cls in classes:
        grouped_data[cls.name] = {}
        cls_subjects = cls.subjects.all()
        
        for section in cls.sections.all():
            grouped_data[cls.name][section.name] = {
                'missing_attendance': [],
                'students': {} # grouped absentees
            }
            
            # Check for missing attendance against all subjects applicable to this section
            for subject in cls_subjects:
                if subject.section_id is not None and subject.section_id != section.id:
                    continue
                
                if (section.id, subject.id) not in marked_sessions:
                    teachers = subject.teachers.all()
                    teacher_names = ", ".join([t.user.get_full_name() or t.user.username for t in teachers]) if teachers else "No Teacher Assigned"
                    grouped_data[cls.name][section.name]['missing_attendance'].append({
                        'subject': subject.name,
                        'teacher': teacher_names
                    })
                    total_missing_attendance += 1
    
    # Get all absent records for the selected date
    absent_records = AttendanceRecord.objects.filter(
        attendance__date=selected_date,
        status='ABSENT'
    ).select_related(
        'student',
        'student__section',
        'student__section__class_ref',
        'attendance__subject',
        'attendance__teacher__user'
    ).order_by(
        'student__section__class_ref__name',
        'student__section__name',
        'student__roll_number'
    )
    
    # Group absentees by class, section, and student
    for record in absent_records:
        class_name = record.student.section.class_ref.name
        section_name = record.student.section.name
        student_id = record.student.id
        
        # In case there are records for a class/section not in the prefetched list (rare but possible)
        if class_name not in grouped_data:
            grouped_data[class_name] = {}
        if section_name not in grouped_data[class_name]:
            grouped_data[class_name][section_name] = {
                'missing_attendance': [],
                'students': {}
            }
            
        if student_id not in grouped_data[class_name][section_name]['students']:
            grouped_data[class_name][section_name]['students'][student_id] = {
                'student': record.student,
                'records': [],
                'manager_remarks': record.manager_remarks or '',
            }
            
        grouped_data[class_name][section_name]['students'][student_id]['records'].append(record)
        
        if record.manager_remarks:
            grouped_data[class_name][section_name]['students'][student_id]['manager_remarks'] = record.manager_remarks
    
    # Calculate total unique absent students
    total_absent_students = absent_records.values('student_id').distinct().count()
    
    context = {
        'selected_date_str': selected_date_str,
        'selected_date': selected_date,
        'grouped_data': grouped_data,
        'total_absent': total_absent_students,
        'total_missing': total_missing_attendance,
    }
    
    return render(request, 'manager/dashboard.html', context)


@manager_required
@require_POST
def save_manager_remarks(request):
    """AJAX endpoint to save manager remarks for a specific student's absence on a specific date"""
    try:
        data = json.loads(request.body)
        student_id = data.get('student_id')
        date_str = data.get('date')
        remarks = data.get('remarks', '')
        
        records = AttendanceRecord.objects.filter(
            student_id=student_id,
            attendance__date=date_str,
            status='ABSENT'
        )
        records.update(manager_remarks=remarks)
        
        return JsonResponse({
            'success': True,
            'message': 'Remarks saved successfully'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@manager_required
def export_manager_report_excel(request):
    """Export manager dashboard data to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    from datetime import date
    
    selected_date_str = request.GET.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        selected_date = date.fromisoformat(selected_date_str)
    except ValueError:
        selected_date = date.today()
        selected_date_str = selected_date.strftime('%Y-%m-%d')
        
    # Data gathering logic from manager_dashboard view
    absent_records = AttendanceRecord.objects.filter(
        attendance__date=selected_date,
        status='ABSENT'
    ).select_related(
        'student',
        'student__section',
        'student__section__class_ref',
        'attendance__subject',
        'attendance__teacher__user'
    ).order_by(
        'student__section__class_ref__name',
        'student__section__name',
        'student__roll_number'
    )
    
    # Export manager dashboard data to Excel
    wb = Workbook()
    
    # --- Sheet 1: Absentee Report ---
    ws_absent = wb.active
    ws_absent.title = "Absentee Report"
    
    ws_absent.merge_cells('A1:G1')
    title_cell = ws_absent['A1']
    title_cell.value = f"Absentee Report - {selected_date.strftime('%b %d, %Y')}"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    headers = ['Roll No', 'Student Name', 'Class', 'Section', 'Subject', 'Teacher', 'Manager Remarks']
    header_fill = PatternFill(start_color='EF4444', end_color='EF4444', fill_type='solid') # Rose color
    header_font = Font(color='FFFFFF', bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_absent.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        
    # Group results by class, section, and student for export
    grouped_records = []
    student_records_map = {}
    
    for record in absent_records:
        student_id = record.student.id
        if student_id not in student_records_map:
            student_records_map[student_id] = {
                'student': record.student,
                'records': [],
                'manager_remarks': record.manager_remarks or ''
            }
            grouped_records.append(student_records_map[student_id])
            
        student_records_map[student_id]['records'].append(record)
        if record.manager_remarks:
            student_records_map[student_id]['manager_remarks'] = record.manager_remarks
            
    row_num = 4
    for data in grouped_records:
        student = data['student']
        subjects = ", ".join([r.attendance.subject.name for r in data['records']])
        teachers = ", ".join([r.attendance.teacher.user.get_full_name() or r.attendance.teacher.user.username for r in data['records']])
        
        ws_absent.cell(row=row_num, column=1, value=student.roll_number).alignment = Alignment(horizontal='center')
        ws_absent.cell(row=row_num, column=2, value=student.name)
        ws_absent.cell(row=row_num, column=3, value=student.section.class_ref.name)
        ws_absent.cell(row=row_num, column=4, value=student.section.name).alignment = Alignment(horizontal='center')
        ws_absent.cell(row=row_num, column=5, value=subjects)
        ws_absent.cell(row=row_num, column=6, value=teachers)
        ws_absent.cell(row=row_num, column=7, value=data['manager_remarks'])
        row_num += 1
        
    if row_num == 4:
        ws_absent.cell(row=4, column=1, value="No absentees found for this date")
        ws_absent.merge_cells('A4:G4')
        ws_absent.cell(row=4, column=1).alignment = Alignment(horizontal='center')
        
    ws_absent.column_dimensions['A'].width = 15
    ws_absent.column_dimensions['B'].width = 25
    ws_absent.column_dimensions['C'].width = 20
    ws_absent.column_dimensions['D'].width = 12
    ws_absent.column_dimensions['E'].width = 25
    ws_absent.column_dimensions['F'].width = 25
    ws_absent.column_dimensions['G'].width = 40
    
    # --- Sheet 2: Missing Attendance Report ---
    ws_missing = wb.create_sheet(title="Missing Attendance")
    
    ws_missing.merge_cells('A1:D1')
    title_cell_missing = ws_missing['A1']
    title_cell_missing.value = f"Missing Attendance Report - {selected_date.strftime('%b %d, %Y')}"
    title_cell_missing.font = Font(size=14, bold=True)
    title_cell_missing.alignment = Alignment(horizontal='center')
    
    headers_missing = ['Class', 'Section', 'Subject', 'Teacher']
    header_fill_missing = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid') # Amber color
    header_font_missing = Font(color='FFFFFF', bold=True)
    
    for col_num, header in enumerate(headers_missing, 1):
        cell = ws_missing.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill_missing
        cell.font = header_font_missing
        cell.alignment = Alignment(horizontal='center')
        
    classes = Class.objects.prefetch_related(
        'sections', 
        'subjects__teachers__user'
    ).all().order_by('name')
    
    attendances_today = Attendance.objects.filter(date=selected_date)
    marked_sessions = set(attendances_today.values_list('section_id', 'subject_id'))
    
    row_num_missing = 4
    for cls in classes:
        cls_subjects = cls.subjects.all()
        for section in cls.sections.all():
            for subject in cls_subjects:
                if subject.section_id is not None and subject.section_id != section.id:
                    continue
                
                if (section.id, subject.id) not in marked_sessions:
                    teachers = subject.teachers.all()
                    teacher_names = ", ".join([t.user.get_full_name() or t.user.username for t in teachers]) if teachers else "No Teacher Assigned"
                    
                    ws_missing.cell(row=row_num_missing, column=1, value=cls.name)
                    ws_missing.cell(row=row_num_missing, column=2, value=section.name).alignment = Alignment(horizontal='center')
                    ws_missing.cell(row=row_num_missing, column=3, value=subject.name)
                    ws_missing.cell(row=row_num_missing, column=4, value=teacher_names)
                    row_num_missing += 1
                    
    if row_num_missing == 4:
        ws_missing.cell(row=4, column=1, value="All attendance records have been marked")
        ws_missing.merge_cells('A4:D4')
        ws_missing.cell(row=4, column=1).alignment = Alignment(horizontal='center')
        
    ws_missing.column_dimensions['A'].width = 20
    ws_missing.column_dimensions['B'].width = 15
    ws_missing.column_dimensions['C'].width = 25
    ws_missing.column_dimensions['D'].width = 30
    
    # Generate Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Manager_Report_{selected_date.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


# Teacher Views
@teacher_required
def teacher_dashboard(request):
    """Teacher dashboard showing assigned classes and subjects"""
    teacher = request.user.teacher_profile
    
    # Get assigned subjects with sections
    assigned_subjects = teacher.assigned_subjects.select_related('class_ref').all()
    
    # Organize subjects with their sections
    subjects_with_sections = []
    for subject in assigned_subjects:
        sections = Section.objects.filter(class_ref=subject.class_ref).order_by('name')
        subjects_with_sections.append({
            'subject': subject,
            'sections': sections
        })
    
    context = {
        'teacher': teacher,
        'assigned_classes': teacher.assigned_classes.all(),
        'assigned_subjects': assigned_subjects,
        'subjects_with_sections': subjects_with_sections,
    }
    return render(request, 'teacher/dashboard.html', context)


@teacher_required
def mark_attendance(request):
    """Mark attendance for students"""
    teacher = request.user.teacher_profile
    students = []
    selected_data = {}
    attendance_exists = False
    
    # Initialize form with teacher's assignments and today's date
    form = AttendanceFilterForm(teacher=teacher, initial={'date': date.today()})
    
    if request.method == 'POST':
        form = AttendanceFilterForm(request.POST, teacher=teacher)
        if form.is_valid():
            class_ref = form.cleaned_data['class_ref']
            section = form.cleaned_data['section']
            subject = form.cleaned_data['subject']
            attendance_date = form.cleaned_data['date']
            
            # Store selected data for template
            selected_data = {
                'class_ref': class_ref,
                'section': section,
                'subject': subject,
                'date': attendance_date,
            }
            
            # Check if attendance already exists
            attendance_qs = Attendance.objects.filter(
                class_ref=class_ref,
                section=section,
                subject=subject,
                date=attendance_date
            )
            attendance_exists = attendance_qs.exists()
            
            # Get students in the selected section and annotate with attendance data
            total_sessions = Attendance.objects.filter(
                class_ref=class_ref,
                section=section,
                subject=subject
            ).count()

            students = Student.objects.filter(section=section).annotate(
                present_count=Count(
                    'attendance_records',
                    filter=Q(
                        attendance_records__attendance__class_ref=class_ref,
                        attendance_records__attendance__section=section,
                        attendance_records__attendance__subject=subject,
                        attendance_records__status='PRESENT'
                    )
                )
            ).order_by('roll_number')

            if attendance_exists:
                messages.info(request, 'Attendance already submitted. View only mode.')
                attendance_obj = attendance_qs.first()
                record_map = {r.student_id: r.status for r in attendance_obj.records.all()}
                for student in students:
                    student.current_status = record_map.get(student.id, 'ABSENT')
            else:
                for student in students:
                    student.current_status = 'PRESENT'
            
            # Calculate percentages
            for student in students:
                if total_sessions > 0:
                    percentage = (student.present_count / total_sessions) * 100
                    student.attendance_percentage = f"{percentage:.1f}"
                else:
                    student.attendance_percentage = "0.0"

            # Check if date is in the past
            is_past_date = attendance_date < timezone.now().date()
    
    context = {
        'form': form,
        'students': students,
        'selected_data': selected_data,
        'attendance_exists': attendance_exists,
        'is_past_date': is_past_date if 'is_past_date' in locals() else False,
    }
    return render(request, 'teacher/mark_attendance.html', context)


@teacher_required
@require_POST
def save_attendance(request):
    """Save attendance records (AJAX endpoint)"""
    try:
        data = json.loads(request.body)
        teacher = request.user.teacher_profile
        
        class_id = data.get('class_id')
        section_id = data.get('section_id')
        subject_id = data.get('subject_id')
        attendance_date = data.get('date')
        records = data.get('records', [])
        
        # Validate data
        class_ref = get_object_or_404(Class, id=class_id)
        section = get_object_or_404(Section, id=section_id)
        subject = get_object_or_404(Subject, id=subject_id)
        
        # Check if attendance already exists
        if Attendance.objects.filter(
            class_ref=class_ref,
            section=section,
            subject=subject,
            date=attendance_date
        ).exists():
            return JsonResponse({
                'success': False,
                'message': 'Attendance already exists for this date.'
            }, status=400)
        
        # Create attendance session
        attendance = Attendance.objects.create(
            class_ref=class_ref,
            section=section,
            subject=subject,
            teacher=teacher,
            date=attendance_date
        )
        
        # Create attendance records for each student
        for record in records:
            student = get_object_or_404(Student, id=record['student_id'])
            AttendanceRecord.objects.create(
                attendance=attendance,
                student=student,
                status=record['status']
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Attendance saved successfully!'
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@login_required
def get_sections(request):
    """AJAX endpoint to get sections for a class"""
    class_id = request.GET.get('class_id')
    print(f"DEBUG: get_sections called with class_id={class_id}")
    if not class_id:
        return JsonResponse([], safe=False)
    
    try:
        sections = Section.objects.filter(class_ref_id=class_id).values('id', 'name')
        print(f"DEBUG: Found {len(sections)} sections")
        return JsonResponse(list(sections), safe=False)
    except Exception as e:
        print(f"DEBUG: Error in get_sections: {e}")
        return JsonResponse([], safe=False)


@login_required
def get_sections_multiple(request):
    """AJAX endpoint to get sections for multiple classes"""
    class_ids = request.GET.getlist('class_ids[]')
    if not class_ids:
        class_ids = request.GET.get('class_ids', '').split(',')
        class_ids = [cid for cid in class_ids if cid.strip()]
        
    print(f"DEBUG: get_sections_multiple called with class_ids={class_ids}")
    if not class_ids:
        return JsonResponse([], safe=False)
    
    try:
        sections = Section.objects.filter(class_ref_id__in=class_ids).select_related('class_ref').order_by('class_ref__name', 'name')
        # Format the data to include class name for better UI display
        data = [{'id': s.id, 'name': f"{s.class_ref.name} - {s.name}"} for s in sections]
        print(f"DEBUG: Found {len(data)} sections")
        return JsonResponse(data, safe=False)
    except Exception as e:
        print(f"DEBUG: Error in get_sections_multiple: {e}")
        return JsonResponse([], safe=False)


@login_required
def get_subjects(request):
    """AJAX endpoint to get subjects for a class (filtered by teacher's assignments if teacher)"""
    class_id = request.GET.get('class_id')
    print(f"DEBUG: get_subjects called with class_id={class_id}")
    if not class_id:
        return JsonResponse([], safe=False)
    
    try:
        # If user is a teacher, filter by their assignments
        if request.user.role == 'TEACHER':
            teacher = request.user.teacher_profile
            subjects = teacher.assigned_subjects.filter(class_ref_id=class_id).values('id', 'name')
        else:
            # If admin, show all subjects for the class
            subjects = Subject.objects.filter(class_ref_id=class_id).values('id', 'name')
        print(f"DEBUG: Found {len(subjects)} subjects")
        return JsonResponse(list(subjects), safe=False)
    except Exception as e:
        print(f"DEBUG: Error in get_subjects: {e}")
        return JsonResponse([], safe=False)


@login_required
def get_subjects_multiple(request):
    """AJAX endpoint to get subjects for multiple classes"""
    class_ids = request.GET.getlist('class_ids[]')
    if not class_ids:
        class_ids = request.GET.get('class_ids', '').split(',')
        class_ids = [cid for cid in class_ids if cid.strip()]
        
    print(f"DEBUG: get_subjects_multiple called with class_ids={class_ids}")
    if not class_ids:
        return JsonResponse([], safe=False)
    
    try:
        if request.user.role == 'TEACHER':
            teacher = request.user.teacher_profile
            subjects = teacher.assigned_subjects.filter(class_ref_id__in=class_ids).select_related('class_ref').order_by('class_ref__name', 'name')
        else:
            subjects = Subject.objects.filter(class_ref_id__in=class_ids).select_related('class_ref').order_by('class_ref__name', 'name')
        
        # Format the data to include class name for better UI display
        data = [{'id': s.id, 'name': f"{s.name} ({s.class_ref.name})"} for s in subjects]
        print(f"DEBUG: Found {len(data)} subjects")
        return JsonResponse(data, safe=False)
    except Exception as e:
        print(f"DEBUG: Error in get_subjects_multiple: {e}")
        return JsonResponse([], safe=False)


@teacher_required
def export_attendance(request):
    """Export attendance data to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    from datetime import datetime, timedelta
    
    teacher = request.user.teacher_profile
    
    if request.method == 'POST':
        form = AttendanceExportForm(request.POST, teacher=teacher)
        if form.is_valid():
            class_ref = form.cleaned_data['class_ref']
            section = form.cleaned_data['section']
            subject = form.cleaned_data['subject']
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            
            # Get attendance records
            attendances = Attendance.objects.filter(
                class_ref=class_ref,
                section=section,
                subject=subject,
                date__range=[start_date, end_date]
            ).prefetch_related('records', 'records__student')
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Report"
            
            # Add title
            ws.merge_cells('A1:F1')
            title_cell = ws['A1']
            title_cell.value = "Attendance Report"
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Add report info
            ws.merge_cells('A2:F2')
            info_cell = ws['A2']
            info_cell.value = f"Class: {class_ref.name} | Section: {section.name} | Subject: {subject.name}"
            info_cell.font = Font(italic=True)
            info_cell.alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A3:F3')
            date_cell = ws['A3']
            date_cell.value = f"Date Range: {start_date} to {end_date}"
            date_cell.alignment = Alignment(horizontal='center')

            ws.merge_cells('A4:F4')
            teacher_cell = ws['A4']
            
            # Using get_full_name() or username for display
            teacher_name = teacher.user.get_full_name() or teacher.user.username
            teacher_cell.value = f"Teacher: {teacher_name}"
            teacher_cell.font = Font(italic=True)
            teacher_cell.alignment = Alignment(horizontal='center')
            
            # Get all dates in range
            delta = end_date - start_date
            dates = [start_date + timedelta(days=i) for i in range(delta.days + 1)]
            
            # Prepare headers
            headers = ['Roll No', 'Name', 'Total Percentage']
            
            # Style headers
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Get all students in section
            students = Student.objects.filter(section=section).order_by('roll_number')
            
            # Organize attendance data
            attendance_map = {} # (student_id, date) -> status
            for att in attendances:
                for record in att.records.all():
                    attendance_map[(record.student_id, att.date)] = record.status
            
            # Write student rows
            for row_num, student in enumerate(students, 6):
                ws.cell(row=row_num, column=1, value=student.roll_number)
                ws.cell(row=row_num, column=2, value=student.name)
                
                present_count = 0
                total_days = 0
                
                for date_obj in dates:
                    status = attendance_map.get((student.id, date_obj), '-')
                    if status == 'PRESENT':
                        present_count += 1
                        total_days += 1
                    elif status == 'ABSENT':
                        total_days += 1
                
                # Percentage
                if total_days > 0:
                    percentage = (present_count / total_days) * 100
                    ws.cell(row=row_num, column=3, value=f"{percentage:.1f}%")
                else:
                     ws.cell(row=row_num, column=3, value="0%")
                
                ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 20
                
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"Attendance_{class_ref.name}_{section.name}_{subject.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
    else:
        form = AttendanceExportForm(teacher=teacher)
    
    return render(request, 'teacher/export_attendance.html', {'form': form})


# ──────────────────────────────────────────────
# Admin Attendance Report Views
# ──────────────────────────────────────────────

@admin_required
def admin_attendance_report(request):
    """Render the attendance report page for admins."""
    classes = Class.objects.all().order_by('name')
    context = {'classes': classes}
    return render(request, 'admin/attendance_report.html', context)


@admin_required
def api_get_report_data(request):
    """
    AJAX endpoint that returns per-student attendance summary
    for the chosen class / section / subject within a date range.
    """
    from datetime import datetime as dt

    class_id   = request.GET.get('class_id')
    section_id = request.GET.get('section_id')
    subject_id = request.GET.get('subject_id')
    start_date = request.GET.get('start_date')
    end_date   = request.GET.get('end_date')

    # ── Validate required params ──
    if not all([class_id, section_id, subject_id, start_date, end_date]):
        return JsonResponse({'error': 'Missing required parameters.'}, status=400)

    try:
        start = dt.strptime(start_date, '%Y-%m-%d').date()
        end   = dt.strptime(end_date,   '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

    if start > end:
        return JsonResponse({'error': 'Start date must be before or equal to end date.'}, status=400)

    # ── Fetch attendance sessions in range ──
    attendances = Attendance.objects.filter(
        class_ref_id=class_id,
        section_id=section_id,
        subject_id=subject_id,
        date__range=[start, end],
    ).prefetch_related('records__student')

    total_sessions = attendances.count()

    # Build a map: student_id → {present, absent}
    student_stats = {}
    student_names = {}
    student_rolls = {}

    for att in attendances:
        for rec in att.records.all():
            sid = rec.student_id
            if sid not in student_stats:
                student_stats[sid] = {'present': 0, 'absent': 0}
                student_names[sid] = rec.student.name
                student_rolls[sid] = rec.student.roll_number
            if rec.status == 'PRESENT':
                student_stats[sid]['present'] += 1
            else:
                student_stats[sid]['absent'] += 1

    # ── Also include students who have no records (100 % absent) ──
    all_students = Student.objects.filter(section_id=section_id).order_by('roll_number')
    for s in all_students:
        if s.id not in student_stats:
            student_stats[s.id]  = {'present': 0, 'absent': 0}
            student_names[s.id]  = s.name
            student_rolls[s.id]  = s.roll_number

    # ── Build result list ──
    rows = []
    for sid, stats in student_stats.items():
        present = stats['present']
        absent  = stats['absent']
        sessions = present + absent if total_sessions == 0 else total_sessions
        pct = round((present / sessions) * 100, 1) if sessions > 0 else 0.0
        rows.append({
            'roll':    student_rolls[sid],
            'name':    student_names[sid],
            'sessions': total_sessions,
            'present': present,
            'absent':  absent,
            'percentage': pct,
        })

    rows.sort(key=lambda r: r['roll'])

    return JsonResponse({
        'total_sessions': total_sessions,
        'rows': rows,
        'start_date': str(start),
        'end_date':   str(end),
    })

@admin_required
def admin_edit_attendance(request):
    """
    Render a page allowing admins to edit attendance for a specific class/section/subject
    within a date range.
    """
    from datetime import datetime as dt

    class_id   = request.GET.get('class_id')
    section_id = request.GET.get('section_id')
    subject_id = request.GET.get('subject_id')
    start_date = request.GET.get('start_date')
    end_date   = request.GET.get('end_date')

    if not all([class_id, section_id, subject_id, start_date, end_date]):
        messages.error(request, "Missing required parameters for editing attendance.")
        return redirect('admin_attendance_report')

    try:
        start = dt.strptime(start_date, '%Y-%m-%d').date()
        end   = dt.strptime(end_date,   '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, "Invalid date format.")
        return redirect('admin_attendance_report')

    try:
        class_obj = Class.objects.get(id=class_id)
        section_obj = Section.objects.get(id=section_id)
        subject_obj = Subject.objects.get(id=subject_id)
    except (Class.DoesNotExist, Section.DoesNotExist, Subject.DoesNotExist):
        messages.error(request, "Invalid class, section, or subject.")
        return redirect('admin_attendance_report')

    # Fetch all attendance sessions in the range for this subject/section
    attendances = Attendance.objects.filter(
        class_ref=class_obj,
        section=section_obj,
        subject=subject_obj,
        date__range=[start, end]
    ).order_by('date').prefetch_related('records', 'records__student')

    # Fetch all students in the section
    students = Student.objects.filter(section=section_obj).order_by('roll_number')

    # Build a matrix: [student][attendance_session.id] -> record.status
    matrix = []
    for student in students:
        student_records = []
        for att in attendances:
            # Find the record for this student/session
            record = next((r for r in att.records.all() if r.student_id == student.id), None)
            student_records.append({
                'attendance_id': att.id,
                'status': record.status if record else 'ABSENT', # Default to ABSENT if no record exists
                'record_id': record.id if record else None
            })
        matrix.append({
            'student': student,
            'records': student_records
        })

    context = {
        'class_obj': class_obj,
        'section_obj': section_obj,
        'subject_obj': subject_obj,
        'start_date': start_date,
        'end_date': end_date,
        'attendances': attendances,
        'matrix': matrix,
        'class_id': class_id,
        'section_id': section_id,
        'subject_id': subject_id,
    }
    return render(request, 'admin/edit_attendance.html', context)

@admin_required
@require_POST
def admin_save_bulk_attendance(request):
    """
    Handle the submission from the edit_attendance.html template.
    Update the attendance records in bulk.
    """
    class_id = request.POST.get('class_id')
    section_id = request.POST.get('section_id')
    subject_id = request.POST.get('subject_id')
    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')

    # Process all keys like status_{attendance_id}_{student_id}
    updates_made = 0
    records_to_create = []
    records_to_update = []

    # Get a mapping of student IDs in this section and attendance IDs in this range to avoid making too many queries
    students_in_section = set(Student.objects.filter(section_id=section_id).values_list('id', flat=True))
    from datetime import datetime as dt
    start = dt.strptime(start_date, '%Y-%m-%d').date()
    end   = dt.strptime(end_date,   '%Y-%m-%d').date()
    attendances_in_range = set(Attendance.objects.filter(
        class_ref_id=class_id,
        section_id=section_id,
        subject_id=subject_id,
        date__range=[start, end]
    ).values_list('id', flat=True))

    for key, value in request.POST.items():
        if key.startswith('status_'):
            try:
                parts = key.split('_')
                if len(parts) == 3:
                    att_id = int(parts[1])
                    student_id = int(parts[2])
                    new_status = value

                    if student_id not in students_in_section or att_id not in attendances_in_range:
                        continue # Security/integrity check

                    if new_status not in ['PRESENT', 'ABSENT']:
                        continue

                    # Find existing record
                    try:
                        record = AttendanceRecord.objects.get(attendance_id=att_id, student_id=student_id)
                        if record.status != new_status:
                            record.status = new_status
                            records_to_update.append(record)
                            updates_made += 1
                    except AttendanceRecord.DoesNotExist:
                        records_to_create.append(AttendanceRecord(
                            attendance_id=att_id,
                            student_id=student_id,
                            status=new_status
                        ))
                        updates_made += 1
            except ValueError:
                pass # Ignore malformed keys

    # Bulk create/update for efficiency
    if records_to_create:
        AttendanceRecord.objects.bulk_create(records_to_create)
    if records_to_update:
        AttendanceRecord.objects.bulk_update(records_to_update, ['status'])

    if updates_made > 0:
        messages.success(request, f"Successfully updated {updates_made} attendance records.")
    else:
        messages.info(request, "No changes were made.")

    from django.urls import reverse
    from urllib.parse import urlencode
    
    # Redirect back to the report page with the same filters
    base_url = reverse('admin_attendance_report')
    query_string = urlencode({
        'class_id': class_id,
        'section_id': section_id,
        'subject_id': subject_id,
        'start_date': start_date,
        'end_date': end_date
    })
    url = f"{base_url}?{query_string}"
    return redirect(url)

@admin_required
@require_POST
def admin_add_attendance_session(request):
    """
    Handle adding a new attendance session from the edit_attendance.html template.
    Creates a new attendance object and default PRESENT records for all students in the section.
    """
    class_id = request.POST.get('class_id')
    section_id = request.POST.get('section_id')
    subject_id = request.POST.get('subject_id')
    session_date = request.POST.get('session_date')
    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')

    from django.urls import reverse
    from urllib.parse import urlencode

    def redirect_back():
        base_url = reverse('admin_edit_attendance')
        query_string = urlencode({
            'class_id': class_id or '',
            'section_id': section_id or '',
            'subject_id': subject_id or '',
            'start_date': start_date or '',
            'end_date': end_date or ''
        })
        return redirect(f"{base_url}?{query_string}")

    if not all([class_id, section_id, subject_id, session_date]):
        messages.error(request, "Missing required parameters.")
        return redirect_back()

    try:
        class_obj = Class.objects.get(id=class_id)
        section_obj = Section.objects.get(id=section_id)
        subject_obj = Subject.objects.get(id=subject_id)
    except (Class.DoesNotExist, Section.DoesNotExist, Subject.DoesNotExist):
        messages.error(request, "Invalid parameters.")
        return redirect_back()

    if Attendance.objects.filter(class_ref=class_obj, section=section_obj, subject=subject_obj, date=session_date).exists():
        messages.error(request, f"Session already exists for {session_date}.")
        return redirect_back()

    teacher = Teacher.objects.filter(assigned_subjects=subject_obj, assigned_sections=section_obj).first()
    if not teacher:
        teacher = Teacher.objects.first()

    if not teacher:
        messages.error(request, "No teacher found in the system to assign to this session.")
        return redirect_back()

    attendance = Attendance.objects.create(
        class_ref=class_obj,
        section=section_obj,
        subject=subject_obj,
        teacher=teacher,
        date=session_date
    )

    students = Student.objects.filter(section=section_obj)
    records = [
        AttendanceRecord(attendance=attendance, student=student, status='PRESENT')
        for student in students
    ]
    AttendanceRecord.objects.bulk_create(records)

    messages.success(request, f"New session created for {session_date}.")
    return redirect_back()
